"""Импорт расписания из Excel (простой табличный формат).

Ожидаемые колонки первой строки (регистр не важен):
Группа | Преподаватель | Дисциплина | Аудитория | День недели | Начало | Конец

Опциональная колонка «Неделя»: белая / зелёная / каждая (пусто = каждая).
День недели — число 1–7 или название («понедельник», «пн», ...).
Недостающие группы/преподаватели/дисциплины/аудитории создаются автоматически.

Институтский формат (сетка с группами по колонкам) обрабатывается
отдельным парсером, см. timetable_import.py.
"""

import io
from datetime import time
from typing import BinaryIO

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session as DbSession

from app.models import Classroom, Discipline, Group, Schedule, Teacher, WeekType
from app.schemas.schedule import ScheduleImportResult

EXPECTED_HEADERS = {
    "группа": "group",
    "преподаватель": "teacher",
    "дисциплина": "discipline",
    "аудитория": "classroom",
    "день недели": "weekday",
    "начало": "starts_at",
    "конец": "ends_at",
}

OPTIONAL_HEADERS = {"неделя": "week_type"}

WEEK_TYPES = {
    "": WeekType.every,
    "каждая": WeekType.every,
    "белая": WeekType.white,
    "зелёная": WeekType.green,
    "зеленая": WeekType.green,
}

WEEKDAYS = {
    "понедельник": 1, "пн": 1,
    "вторник": 2, "вт": 2,
    "среда": 3, "ср": 3,
    "четверг": 4, "чт": 4,
    "пятница": 5, "пт": 5,
    "суббота": 6, "сб": 6,
    "воскресенье": 7, "вс": 7,
}


def build_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Расписание"
    ws.append(
        ["Группа", "Преподаватель", "Дисциплина", "Аудитория", "День недели", "Начало", "Конец", "Неделя"]
    )
    ws.append(["ИС-31", "Иванов И.И.", "Базы данных", "301", "понедельник", "09:00", "10:30", "каждая"])
    ws.append(["ИС-31", "Петров П.П.", "Физика", "409", "вторник", "13:20", "14:50", "белая"])
    for column in "ABCDEFGH":
        ws.column_dimensions[column].width = 20
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _parse_weekday(value) -> int:
    if isinstance(value, (int, float)):
        weekday = int(value)
        if 1 <= weekday <= 7:
            return weekday
        raise ValueError(f"день недели должен быть 1–7, получено {weekday}")
    name = str(value).strip().lower()
    if name in WEEKDAYS:
        return WEEKDAYS[name]
    raise ValueError(f"не удалось распознать день недели «{value}»")


def _parse_time(value) -> time:
    if isinstance(value, time):
        return value
    text = str(value).strip()
    for fmt in ("%H:%M", "%H:%M:%S", "%H.%M"):
        try:
            from datetime import datetime

            return datetime.strptime(text, fmt).time()
        except ValueError:
            continue
    raise ValueError(f"не удалось распознать время «{value}»")


def _map_headers(header_row) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, cell in enumerate(header_row):
        if cell is None:
            continue
        title = str(cell).strip().lower()
        key = EXPECTED_HEADERS.get(title) or OPTIONAL_HEADERS.get(title)
        if key:
            mapping[key] = index
    missing = set(EXPECTED_HEADERS.values()) - set(mapping)
    if missing:
        raise ValueError(
            "в файле не найдены колонки: "
            + ", ".join(sorted(h for h, k in EXPECTED_HEADERS.items() if k in missing))
        )
    return mapping


def _get_or_create(db: DbSession, model, filter_by: dict, defaults: dict | None = None):
    instance = db.scalars(select(model).filter_by(**filter_by)).one_or_none()
    if instance is None:
        instance = model(**filter_by, **(defaults or {}))
        db.add(instance)
        db.flush()
    return instance


def import_schedule(db: DbSession, file: BinaryIO) -> ScheduleImportResult:
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    try:
        header = next(rows)
    except StopIteration:
        return ScheduleImportResult(created=0, skipped=0, errors=["файл пуст"])

    columns = _map_headers(header)
    created = skipped = 0
    errors: list[str] = []

    for row_number, row in enumerate(rows, start=2):
        if row is None or all(cell is None for cell in row):
            continue
        try:
            group_name = str(row[columns["group"]]).strip()
            teacher_name = str(row[columns["teacher"]]).strip()
            discipline_name = str(row[columns["discipline"]]).strip()
            classroom_number = str(row[columns["classroom"]]).strip()
            weekday = _parse_weekday(row[columns["weekday"]])
            starts_at = _parse_time(row[columns["starts_at"]])
            ends_at = _parse_time(row[columns["ends_at"]])
            if ends_at <= starts_at:
                raise ValueError("время окончания раньше времени начала")

            week_type = WeekType.every
            if "week_type" in columns:
                raw_week = str(row[columns["week_type"]] or "").strip().lower()
                if raw_week not in WEEK_TYPES:
                    raise ValueError(f"не удалось распознать неделю «{raw_week}»")
                week_type = WEEK_TYPES[raw_week]

            group = _get_or_create(db, Group, {"name": group_name})
            teacher = _get_or_create(db, Teacher, {"full_name": teacher_name})
            discipline = _get_or_create(db, Discipline, {"name": discipline_name})
            classroom = _get_or_create(db, Classroom, {"number": classroom_number})

            duplicate = db.scalars(
                select(Schedule).where(
                    Schedule.group_id == group.id,
                    Schedule.weekday == weekday,
                    Schedule.starts_at == starts_at,
                    Schedule.week_type == week_type,
                )
            ).one_or_none()
            if duplicate:
                skipped += 1
                continue

            db.add(
                Schedule(
                    group_id=group.id,
                    teacher_id=teacher.id,
                    discipline_id=discipline.id,
                    classroom_id=classroom.id,
                    weekday=weekday,
                    starts_at=starts_at,
                    ends_at=ends_at,
                    week_type=week_type,
                )
            )
            db.flush()
            created += 1
        except (ValueError, IndexError, TypeError) as exc:
            errors.append(f"строка {row_number}: {exc}")

    db.commit()
    return ScheduleImportResult(created=created, skipped=skipped, errors=errors)
