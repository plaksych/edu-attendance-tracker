"""Импорт институтского расписания (формат УМО, конвертация из PDF).

Особенности формата:
- один семестр разбит на десятки листов «Table N»; блок листов одного курса
  начинается с баннера «РАСПИСАНИЕ ЗАНЯТИЙ N КУРСА ...»;
- группы идут по колонкам, у каждой — своя колонка «Аудитория»;
- строка с временем пары («13:20-14:50») задаёт слот; слот занимает две строки
  (объединение в колонке времени) либо одну;
- в двухстрочном слоте ячейка на обе строки — пара каждую неделю,
  только верхняя строка — белая неделя, только нижняя — зелёная;
- в ячейке пары: «Дисциплина [звание Фамилия И.О.]», в колонке аудитории:
  «номер [лек.|пр.|лаб.]».
"""

import logging
import re
from dataclasses import dataclass
from datetime import time
from typing import BinaryIO

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session as DbSession

from app.models import Classroom, Discipline, Group, Schedule, Teacher, WeekType
from app.schemas.schedule import ScheduleImportResult

logger = logging.getLogger(__name__)

DAY_NAMES = {
    "понедельник": 1,
    "вторник": 2,
    "среда": 3,
    "четверг": 4,
    "пятница": 5,
    "суббота": 6,
    "воскресенье": 7,
}

TIME_RE = re.compile(r"(\d{1,2})[:.](\d{2})\s*[-–—]\s*(\d{1,2})[:.](\d{2})")
COURSE_RE = re.compile(r"(\d)\s*КУРСА", re.IGNORECASE)
# «звание Фамилия И.О.» в конце ячейки; звания — как в исходном расписании
TEACHER_RE = re.compile(
    r"\b((?:ст\.\s*пр|проф|доц|асс|пр)\.\s*[А-ЯЁ][а-яёА-ЯЁ-]+\s+[А-ЯЁ]\.\s*[А-ЯЁ]\.)"
)
# Запасной вариант: «Фамилия И.О.» в конце без звания
TEACHER_FALLBACK_RE = re.compile(r"([А-ЯЁ][а-яё-]+\s+[А-ЯЁ]\.\s*[А-ЯЁ]\.)\s*$")
LESSON_TYPES = {"лек", "пр", "лаб"}


def looks_like_timetable(file: BinaryIO) -> bool:
    """Определяет институтский формат по заголовку «День недели» на первом листе."""
    file.seek(0)
    wb = load_workbook(file, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=1, max_row=5, max_col=3, values_only=True):
            if any(c and str(c).strip().lower().startswith("день недели") for c in row):
                return True
        return False
    finally:
        wb.close()
        file.seek(0)


@dataclass
class ParsedLesson:
    group: str
    course: int
    weekday: int
    starts_at: time
    ends_at: time
    week_type: WeekType
    discipline: str
    teacher: str | None
    classroom: str | None
    lesson_type: str | None


class _Sheet:
    """Обёртка над листом: значения с учётом объединённых ячеек."""

    def __init__(self, ws: Worksheet) -> None:
        self.ws = ws
        # Для каждой ячейки объединённой области — её верхняя строка и значение
        self._merge_top: dict[tuple[int, int], tuple[int, int]] = {}
        for rng in ws.merged_cells.ranges:
            for row in range(rng.min_row, rng.max_row + 1):
                for col in range(rng.min_col, rng.max_col + 1):
                    self._merge_top[(row, col)] = (rng.min_row, rng.min_col)

    def value(self, row: int, col: int):
        top = self._merge_top.get((row, col), (row, col))
        return self.ws.cell(*top).value

    def cell_rows(self, row: int, col: int) -> tuple[int, int]:
        """Диапазон строк, который занимает ячейка (с учётом объединения)."""
        for rng in self.ws.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                return rng.min_row, rng.max_row
        return row, row

    def text(self, row: int, col: int) -> str:
        value = self.value(row, col)
        if value is None:
            return ""
        return " ".join(str(value).replace("\xa0", " ").split())


def _find_header_row(sheet: _Sheet) -> int | None:
    for row in range(1, min(sheet.ws.max_row, 6) + 1):
        if sheet.text(row, 1).lower().startswith("день недели"):
            return row
    return None


def _group_columns(sheet: _Sheet, header_row: int) -> list[tuple[str, int, int]]:
    """[(имя группы, колонка группы, колонка аудитории)]"""
    headers: dict[int, str] = {}
    for col in range(1, sheet.ws.max_column + 1):
        headers[col] = sheet.text(header_row, col)

    result = []
    group_cols = [
        (col, text.removeprefix("Группа").strip())
        for col, text in headers.items()
        if text.startswith("Группа")
    ]
    for col, name in group_cols:
        room_col = next(
            (c for c in range(col + 1, sheet.ws.max_column + 1) if "Аудито" in headers[c]),
            None,
        )
        if room_col is None:
            logger.warning("Лист %s: у группы %s нет колонки аудитории", sheet.ws.title, name)
            continue
        result.append((name, col, room_col))
    return result


def _parse_time_range(text: str) -> tuple[time, time] | None:
    match = TIME_RE.search(text)
    if not match:
        return None
    h1, m1, h2, m2 = (int(g) for g in match.groups())
    return time(h1, m1), time(h2, m2)


def _split_lesson_text(text: str) -> tuple[str, str | None]:
    """«Физика ст.пр. Строковская С.Е.» -> («Физика», «ст.пр. Строковская С.Е.»)"""
    match = TEACHER_RE.search(text) or TEACHER_FALLBACK_RE.search(text)
    if not match:
        return text.strip(" ,;"), None
    discipline = text[: match.start()].strip(" ,;")
    teacher = text[match.start() :].strip(" ,;")
    return discipline, teacher


def _split_room_text(text: str) -> tuple[str | None, str | None]:
    """«9-406 лаб.» -> («9-406», «лаб.»); «спортзал пр.» -> («спортзал», «пр.»)"""
    if not text:
        return None, None
    # В конвертированных из PDF файлах встречаются мусорные слэши-переносы
    tokens = [token.strip("/") for token in text.split()]
    tokens = [token for token in tokens if token]
    if tokens and tokens[-1].rstrip(".").lower() in LESSON_TYPES:
        lesson_type = tokens[-1].rstrip(".") + "."
        room = " ".join(tokens[:-1])
        return room or None, lesson_type
    return text, None


def _slot_row_spans(sheet: _Sheet, header_row: int) -> list[tuple[int, int, int]]:
    """Слоты листа: [(строка начала, строка конца, weekday)].

    Слот начинается строкой со временем в колонке B; его высота — высота
    объединения в B (либо одна строка). День тянется по колонке A.
    """
    spans = []
    weekday: int | None = None
    row = header_row + 1
    while row <= sheet.ws.max_row:
        day_text = sheet.text(row, 1).lower()
        if day_text:
            if day_text in DAY_NAMES:
                weekday = DAY_NAMES[day_text]
            elif not day_text.startswith("день недели"):
                # Колонтитул («Места осуществления…», подписи) — конец данных
                break
        if weekday and _parse_time_range(sheet.text(row, 2)):
            top, bottom = sheet.cell_rows(row, 2)
            if top == row:
                spans.append((top, bottom, weekday))
                row = bottom + 1
                continue
        row += 1
    return spans


def parse_workbook(file: BinaryIO) -> tuple[list[ParsedLesson], list[str]]:
    wb = load_workbook(file, data_only=True)
    lessons: list[ParsedLesson] = []
    errors: list[str] = []
    course = 1

    for sheet_name in wb.sheetnames:
        sheet = _Sheet(wb[sheet_name])
        header_row = _find_header_row(sheet)
        if header_row is None:
            errors.append(f"{sheet_name}: не найдена строка заголовка, лист пропущен")
            continue

        # Баннер «РАСПИСАНИЕ ЗАНЯТИЙ N КУРСА» открывает блок листов курса
        for row in range(1, header_row):
            for col in range(1, sheet.ws.max_column + 1):
                match = COURSE_RE.search(sheet.text(row, col))
                if match:
                    course = int(match.group(1))

        groups = _group_columns(sheet, header_row)
        if not groups:
            errors.append(f"{sheet_name}: не найдены колонки групп, лист пропущен")
            continue

        for slot_top, slot_bottom, weekday in _slot_row_spans(sheet, header_row):
            time_range = _parse_time_range(sheet.text(slot_top, 2))
            if time_range is None:
                continue
            starts_at, ends_at = time_range

            for group_name, group_col, room_col in groups:
                seen_tops: set[int] = set()
                for row in range(slot_top, slot_bottom + 1):
                    text = sheet.text(row, group_col)
                    if not text:
                        continue
                    cell_top, cell_bottom = sheet.cell_rows(row, group_col)
                    if cell_top in seen_tops:
                        continue
                    seen_tops.add(cell_top)

                    if slot_bottom == slot_top or (
                        cell_top <= slot_top and cell_bottom >= slot_bottom
                    ):
                        week_type = WeekType.every
                    elif cell_top == slot_top:
                        week_type = WeekType.white
                    else:
                        week_type = WeekType.green

                    discipline, teacher = _split_lesson_text(text)
                    room, lesson_type = _split_room_text(sheet.text(cell_top, room_col))
                    if not discipline:
                        continue
                    lessons.append(
                        ParsedLesson(
                            group=group_name,
                            course=course,
                            weekday=weekday,
                            starts_at=starts_at,
                            ends_at=ends_at,
                            week_type=week_type,
                            discipline=discipline,
                            teacher=teacher,
                            classroom=room,
                            lesson_type=lesson_type,
                        )
                    )
    return lessons, errors


def _get_or_create(db: DbSession, cache: dict, model, filter_by: dict, defaults: dict | None = None):
    key = (model, tuple(sorted(filter_by.items())))
    if key in cache:
        return cache[key]
    instance = db.scalars(select(model).filter_by(**filter_by)).one_or_none()
    if instance is None:
        instance = model(**filter_by, **(defaults or {}))
        db.add(instance)
        db.flush()
    cache[key] = instance
    return instance


def import_timetable(db: DbSession, file: BinaryIO) -> ScheduleImportResult:
    lessons, errors = parse_workbook(file)
    cache: dict = {}
    created = skipped = 0

    existing = {
        (s.group_id, s.weekday, s.starts_at, s.week_type)
        for s in db.scalars(select(Schedule)).all()
    }

    for lesson in lessons:
        group = _get_or_create(
            db, cache, Group, {"name": lesson.group}, {"course": lesson.course}
        )
        discipline = _get_or_create(db, cache, Discipline, {"name": lesson.discipline})
        teacher = (
            _get_or_create(db, cache, Teacher, {"full_name": lesson.teacher})
            if lesson.teacher
            else None
        )
        classroom = (
            _get_or_create(db, cache, Classroom, {"number": lesson.classroom})
            if lesson.classroom
            else None
        )

        key = (group.id, lesson.weekday, lesson.starts_at, lesson.week_type)
        if key in existing:
            skipped += 1
            continue
        existing.add(key)

        db.add(
            Schedule(
                group_id=group.id,
                teacher_id=teacher.id if teacher else None,
                discipline_id=discipline.id,
                classroom_id=classroom.id if classroom else None,
                weekday=lesson.weekday,
                starts_at=lesson.starts_at,
                ends_at=lesson.ends_at,
                week_type=lesson.week_type,
                lesson_type=lesson.lesson_type,
            )
        )
        created += 1

    db.commit()
    return ScheduleImportResult(created=created, skipped=skipped, errors=errors)
